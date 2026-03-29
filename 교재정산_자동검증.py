import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

st.set_page_config(page_title="교재 정산 자동 검증", page_icon="📚", layout="wide")

st.markdown("""
<style>
    .stApp { background-color: #f8f9fa; }
    .upload-box {
        background: white;
        border: 2px dashed #4a90d9;
        border-radius: 16px;
        padding: 20px;
        text-align: center;
        min-height: 180px;
    }
    .box-title {
        font-size: 18px;
        font-weight: bold;
        color: #333;
        margin-bottom: 8px;
    }
    .box-desc {
        font-size: 13px;
        color: #888;
        margin-bottom: 12px;
    }
    .result-ok { color: #28a745; font-weight: bold; font-size: 18px; }
    .result-err { color: #dc3545; font-weight: bold; font-size: 18px; }
    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        text-align: center;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 style='text-align:center;'>📚 교재 정산 자동 검증 시스템</h1>", unsafe_allow_html=True)
st.markdown("<p style='text-align:center; color:#666;'>파일 3개를 업로드하면 자동으로 교차 검증합니다</p>", unsafe_allow_html=True)
st.markdown("---")

col1, col2, col3 = st.columns(3, gap="large")

with col1:
    st.markdown('<div class="upload-box">', unsafe_allow_html=True)
    st.markdown('<div class="box-title">1. 계정별원장</div>', unsafe_allow_html=True)
    st.markdown('<div class="box-desc">입고 금액 / 매출 금액 확인용</div>', unsafe_allow_html=True)
    file1 = st.file_uploader("파일 선택", type=["xlsx", "xls"], key="file1", label_visibility="collapsed")
    if file1:
        st.success(f"✅ {file1.name}")
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="upload-box">', unsafe_allow_html=True)
    st.markdown('<div class="box-title">2. 교재정산서</div>', unsafe_allow_html=True)
    st.markdown('<div class="box-desc">관별 재고/판매/입고 실사표</div>', unsafe_allow_html=True)
    file2 = st.file_uploader("파일 선택", type=["xlsx", "xls"], key="file2", label_visibility="collapsed")
    if file2:
        st.success(f"✅ {file2.name}")
    st.markdown('</div>', unsafe_allow_html=True)

with col3:
    st.markdown('<div class="upload-box">', unsafe_allow_html=True)
    st.markdown('<div class="box-title">3. 수익코드 분배</div>', unsafe_allow_html=True)
    st.markdown('<div class="box-desc">관별 수익코드 매핑 파일</div>', unsafe_allow_html=True)
    file3 = st.file_uploader("파일 선택", type=["xlsx", "xls"], key="file3", label_visibility="collapsed")
    if file3:
        st.success(f"✅ {file3.name}")
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown("")

def find_정산시트(wb):
    """본사수정본 시트를 구조로 찾기"""
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
            for cell in row:
                if cell and "재고자산 실사표" in str(cell):
                    if "수정본" in name:
                        return name, ws
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
            for cell in row:
                if cell and "재고자산 실사표" in str(cell):
                    return name, ws
    return None, None

def parse_정산서(ws):
    """정산서 시트에서 교재 데이터 추출"""
    data = []
    campus_code = ""
    for row in ws.iter_rows(min_row=4, max_row=6, max_col=10, values_only=False):
        for cell in row:
            if cell.value and "캠퍼스" in str(cell.value):
                r = cell.row
                campus_cell = ws.cell(row=r, column=4)
                if campus_cell.value:
                    campus_code = str(campus_cell.value).strip()

    header_row = None
    for row in ws.iter_rows(min_row=8, max_row=12, max_col=30, values_only=False):
        for cell in row:
            if cell.value and "No." in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        return campus_code, pd.DataFrame()

    data_start = header_row + 3
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        vals = {cell.column: cell.value for cell in row if cell.value is not None}
        if not vals or len(vals) < 3:
            continue

        row_data = {}
        for cell in row:
            col = cell.column
            v = cell.value
            if col == 2: row_data["No"] = v
            elif col == 3: row_data["캠퍼스"] = v
            elif col == 11: row_data["교재명"] = v
            elif col == 12: row_data["전월재고수량"] = v
            elif col == 13: row_data["매입단가"] = v
            elif col == 14: row_data["전월재고금액"] = v
            elif col == 15: row_data["판매수량"] = v
            elif col == 16: row_data["판매가"] = v
            elif col == 17: row_data["판매금액"] = v
            elif col == 18: row_data["입고수량"] = v
            elif col == 19: row_data["입고단가"] = v
            elif col == 20: row_data["입고금액"] = v

        if "교재명" in row_data and row_data["교재명"]:
            if "ex.)" not in str(row_data.get("No", "")):
                data.append(row_data)

    return campus_code, pd.DataFrame(data)


# 분석 버튼
if file2:
    st.markdown("")
    col_btn = st.columns([1, 1, 1])
    with col_btn[1]:
        analyze = st.button("🔍 분석 시작", use_container_width=True, type="primary")

    if analyze:
        st.markdown("---")
        st.markdown("## 📊 분석 결과")

        try:
            wb = openpyxl.load_workbook(file2, data_only=True)
            시트명, ws = find_정산시트(wb)

            if ws is None:
                st.error("정산서 시트를 찾을 수 없습니다.")
            else:
                st.info(f"📋 분석 시트: **{시트명}**")
                campus, df = parse_정산서(ws)

                if campus:
                    st.markdown(f"### 🏫 캠퍼스: {campus}")

                if not df.empty:
                    numeric_cols = ["전월재고수량", "매입단가", "전월재고금액", "판매수량", "판매가", "판매금액", "입고수량", "입고단가", "입고금액"]
                    for c in numeric_cols:
                        if c in df.columns:
                            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

                    m1, m2, m3, m4 = st.columns(4)
                    with m1:
                        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                        st.metric("교재 종류", f"{len(df)}종")
                        st.markdown('</div>', unsafe_allow_html=True)
                    with m2:
                        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                        total_sale = int(df["판매금액"].sum()) if "판매금액" in df.columns else 0
                        st.metric("총 판매금액", f"{total_sale:,}원")
                        st.markdown('</div>', unsafe_allow_html=True)
                    with m3:
                        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                        total_input = int(df["입고금액"].sum()) if "입고금액" in df.columns else 0
                        st.metric("총 입고금액", f"{total_input:,}원")
                        st.markdown('</div>', unsafe_allow_html=True)
                    with m4:
                        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                        total_stock = int(df["전월재고금액"].sum()) if "전월재고금액" in df.columns else 0
                        st.metric("총 재고금액", f"{total_stock:,}원")
                        st.markdown('</div>', unsafe_allow_html=True)

                    st.markdown("")
                    st.markdown("### 📋 교재별 상세 내역")
                    display_cols = [c for c in ["교재명", "전월재고수량", "매입단가", "판매수량", "판매가", "판매금액", "입고수량", "입고금액"] if c in df.columns]
                    st.dataframe(
                        df[display_cols].style.format({
                            c: "{:,.0f}" for c in display_cols if c != "교재명"
                        }),
                        use_container_width=True,
                        hide_index=True
                    )

                    if file1:
                        st.markdown("### 🔄 계정별원장 vs 정산서 교차검증")
                        st.info("계정별원장 파일의 구조를 파악한 후 자동 대조가 가능합니다. (파일 구조 확인 후 업데이트 예정)")

                    if file3:
                        st.markdown("### 🔄 수익코드 분배 vs 정산서 교차검증")
                        st.info("수익코드 분배 파일의 구조를 파악한 후 자동 대조가 가능합니다. (파일 구조 확인 후 업데이트 예정)")

                    if not file1 and not file3:
                        st.markdown("")
                        st.warning("💡 계정별원장과 수익코드 분배 파일도 업로드하면 교차 검증이 가능합니다!")

                else:
                    st.warning("교재 데이터를 추출하지 못했습니다. 시트 구조를 확인해주세요.")

        except Exception as e:
            st.error(f"파일 분석 중 오류: {str(e)}")

else:
    st.markdown("")
    st.markdown("<p style='text-align:center; color:#999; font-size:16px; margin-top:40px;'>⬆️ 교재정산서 파일을 먼저 업로드해주세요</p>", unsafe_allow_html=True)

st.markdown("---")
st.markdown("<p style='text-align:center; color:#aaa; font-size:12px;'>교재 정산 자동 검증 시스템 v1.0</p>", unsafe_allow_html=True)
