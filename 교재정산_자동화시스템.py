import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from copy import copy

st.set_page_config(page_title="교재 정산 자동화 시스템", page_icon="📚", layout="wide")

st.markdown("""
<style>
.stApp { background: #f5f7fa; }
.main-title { text-align:center; color:#1a73e8; font-size:28px; font-weight:700; margin-bottom:4px; }
.sub-title { text-align:center; color:#666; font-size:14px; margin-bottom:24px; }
.step-header { background:linear-gradient(135deg,#1a73e8,#4285f4); color:white; padding:12px 20px; border-radius:10px; font-size:16px; font-weight:600; margin:16px 0 12px 0; }
.metric-box { background:white; border-radius:10px; padding:16px; box-shadow:0 2px 6px rgba(0,0,0,0.06); text-align:center; }
.metric-num { font-size:24px; font-weight:700; color:#1a73e8; }
.metric-label { font-size:12px; color:#888; }
.warn-box { background:#fff3e0; border-left:4px solid #ff9800; padding:12px 16px; border-radius:0 8px 8px 0; margin:8px 0; font-size:13px; }
.err-box { background:#fce8e6; border-left:4px solid #ea4335; padding:12px 16px; border-radius:0 8px 8px 0; margin:8px 0; font-size:13px; }
.ok-box { background:#e6f4ea; border-left:4px solid #34a853; padding:12px 16px; border-radius:0 8px 8px 0; margin:8px 0; font-size:13px; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">📚 교재 정산 자동화 시스템</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">FLOW 데이터에서 판매 내역을 자동으로 추출하여 정산서를 채워줍니다</div>', unsafe_allow_html=True)

# ── STEP 1: 파일 업로드 ──
st.markdown('<div class="step-header">📁 STEP 1. 파일 업로드</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown("**교재 정산 파일** (FLOW 데이터 + 정산서 포함)")
    uploaded = st.file_uploader("정산 파일 선택", type=["xlsx","xls"], key="main_file", label_visibility="collapsed")

with col2:
    st.markdown("**캠퍼스 코드 입력** (수익코드 필터용)")
    campus_input = st.text_input("예: 수지_203", value="수지_203", label_visibility="collapsed")

if not uploaded:
    st.info("⬆️ 교재 정산 파일을 업로드해주세요")
    st.stop()

# ── 파일 로드 ──
@st.cache_data
def load_file(file_bytes, filename):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    return wb.sheetnames

file_bytes = uploaded.read()
sheet_names = load_file(file_bytes, uploaded.name)
st.success(f"✅ {uploaded.name} 로드 완료 — 시트 {len(sheet_names)}개")

# ── STEP 2: 시트 선택 ──
st.markdown('<div class="step-header">📋 STEP 2. 시트 선택</div>', unsafe_allow_html=True)

col_a, col_b = st.columns(2)
with col_a:
    flow_sheets = [s for s in sheet_names if 'FLOW' in s.upper() or '매출' in s]
    flow_sheet = st.selectbox("FLOW 수납 시트", flow_sheets if flow_sheets else sheet_names)
with col_b:
    jeongsan_sheets = [s for s in sheet_names if '수정본' in s or '실사' in s or (campus_input.split('_')[0] if '_' in campus_input else '') in s]
    if not jeongsan_sheets:
        jeongsan_sheets = [s for s in sheet_names if s not in flow_sheets and 'Log' not in s and 'MF' not in s and '작성' not in s and '피벗' not in s]
    jeongsan_sheet = st.selectbox("정산서 시트 (본사수정본)", jeongsan_sheets if jeongsan_sheets else sheet_names)

# ── STEP 3: 분석 ──
if st.button("🚀 자동 분석 시작", type="primary", use_container_width=True):
    st.markdown('<div class="step-header">📊 STEP 3. 분석 결과</div>', unsafe_allow_html=True)

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    # ── FLOW 데이터 읽기 ──
    ws_flow = wb[flow_sheet]
    flow_rows = []
    headers = []
    for i, row in enumerate(ws_flow.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) if h else f"col{j}" for j, h in enumerate(row)]
            continue
        flow_rows.append(row)

    df_flow = pd.DataFrame(flow_rows, columns=headers[:len(flow_rows[0])] if flow_rows else headers)

    # 수익코드 필터
    code_col = [c for c in df_flow.columns if '수익코드' in str(c)]
    if code_col:
        code_col = code_col[0]
        filtered = df_flow[df_flow[code_col].astype(str).str.contains(campus_input, na=False)].copy()
    else:
        st.error("수익코드 컬럼을 찾을 수 없습니다")
        st.stop()

    # 금액 컬럼
    amt_col = [c for c in df_flow.columns if '납입금액' in str(c)]
    amt_col = amt_col[0] if amt_col else None
    name_col = [c for c in df_flow.columns if '교재명' in str(c) or '반명' in str(c)]
    name_col = name_col[0] if name_col else None

    if not amt_col or not name_col:
        st.error("납입금액 또는 교재명 컬럼을 찾을 수 없습니다")
        st.stop()

    filtered[amt_col] = pd.to_numeric(filtered[amt_col], errors='coerce').fillna(0)

    # 교재명 정리
    filtered['교재명_정리'] = filtered[name_col].astype(str).apply(
        lambda x: x.split(']')[-1].strip() if ']' in x else x
    )

    # A코드(자체교재) vs C코드(시판교재) 분리
    filtered_a = filtered[filtered[code_col].astype(str).str.contains('_A')]
    filtered_c = filtered[filtered[code_col].astype(str).str.contains('_C')]

    # ── 요약 메트릭 ──
    total_all = filtered[amt_col].sum()
    total_a = filtered_a[amt_col].sum()
    total_c = filtered_c[amt_col].sum()
    total_count = len(filtered)
    refund_count = len(filtered[filtered[amt_col] < 0])

    m1, m2, m3, m4, m5 = st.columns(5)
    with m1:
        st.markdown(f'<div class="metric-box"><div class="metric-label">전체 건수</div><div class="metric-num">{total_count:,}</div></div>', unsafe_allow_html=True)
    with m2:
        st.markdown(f'<div class="metric-box"><div class="metric-label">총 납입금액</div><div class="metric-num">{total_all:,.0f}</div></div>', unsafe_allow_html=True)
    with m3:
        st.markdown(f'<div class="metric-box"><div class="metric-label">자체교재(A)</div><div class="metric-num">{total_a:,.0f}</div></div>', unsafe_allow_html=True)
    with m4:
        st.markdown(f'<div class="metric-box"><div class="metric-label">시판교재(C)</div><div class="metric-num">{total_c:,.0f}</div></div>', unsafe_allow_html=True)
    with m5:
        st.markdown(f'<div class="metric-box"><div class="metric-label">환불 건수</div><div class="metric-num" style="color:#ea4335">{refund_count}</div></div>', unsafe_allow_html=True)

    st.markdown("")

    # ── 교재별 판매 집계 ──
    tab1, tab2, tab3 = st.tabs(["📗 자체교재 (A코드)", "📘 시판교재 (C코드)", "📊 정산서 대조"])

    with tab1:
        if len(filtered_a) > 0:
            a_summary = filtered_a.groupby('교재명_정리').agg(
                판매건수=(amt_col, 'count'),
                총납입금액=(amt_col, 'sum'),
                환불건수=(amt_col, lambda x: (x < 0).sum()),
                순매출=(amt_col, lambda x: x[x > 0].sum()),
            ).sort_values('총납입금액', ascending=False).reset_index()
            a_summary.columns = ['교재명', '판매건수', '총납입금액', '환불건수', '순매출']
            st.dataframe(a_summary.style.format({
                '총납입금액': '{:,.0f}', '순매출': '{:,.0f}'
            }), use_container_width=True, hide_index=True)

            # 풀이노트 합계
            note_rows = a_summary[a_summary['교재명'].str.contains('풀이노트', na=False)]
            if len(note_rows) > 0:
                st.markdown(f'<div class="ok-box">📝 <b>풀이노트 합계</b>: {note_rows["판매건수"].sum()}건 / {note_rows["총납입금액"].sum():,.0f}원</div>', unsafe_allow_html=True)
        else:
            st.info("자체교재(A코드) 데이터가 없습니다")

    with tab2:
        if len(filtered_c) > 0:
            c_summary = filtered_c.groupby('교재명_정리').agg(
                판매건수=(amt_col, 'count'),
                총납입금액=(amt_col, 'sum'),
            ).sort_values('총납입금액', ascending=False).reset_index()
            c_summary.columns = ['교재명', '판매건수', '총납입금액']
            st.dataframe(c_summary.style.format({
                '총납입금액': '{:,.0f}'
            }), use_container_width=True, hide_index=True)
        else:
            st.info("시판교재(C코드) 데이터가 없습니다")

    with tab3:
        st.markdown("#### FLOW vs 정산서 대조")

        # 정산서 읽기
        ws_js = wb[jeongsan_sheet]
        js_data = []
        for row in ws_js.iter_rows(min_row=13, max_row=ws_js.max_row, max_col=30, values_only=False):
            rd = {}
            for cell in row:
                c = cell.column
                v = cell.value
                if c == 11: rd['교재명'] = v
                if c == 15: rd['정산서_판매수량'] = v
                if c == 17: rd['정산서_판매금액'] = v
            if rd.get('교재명') and 'ex.)' not in str(rd.get('교재명', '')):
                for k in ['정산서_판매수량', '정산서_판매금액']:
                    try: rd[k] = float(rd.get(k, 0) or 0)
                    except: rd[k] = 0
                js_data.append(rd)

        if js_data:
            df_js = pd.DataFrame(js_data)
            df_js = df_js[df_js['정산서_판매금액'] > 0]

            # FLOW 합계
            flow_total_a = total_a
            js_total = df_js['정산서_판매금액'].sum()

            col_x, col_y, col_z = st.columns(3)
            with col_x:
                st.metric("FLOW 자체교재 합계", f"{flow_total_a:,.0f}원")
            with col_y:
                st.metric("정산서 판매금액 합계", f"{js_total:,.0f}원")
            with col_z:
                diff = flow_total_a - js_total
                st.metric("차이", f"{diff:,.0f}원", delta=f"{diff:,.0f}")

            if abs(diff) < 1:
                st.markdown('<div class="ok-box">✅ <b>FLOW 합계와 정산서 합계가 일치합니다!</b></div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="err-box">⚠️ <b>차이 발생: {diff:,.0f}원</b> — 항목별 확인이 필요합니다</div>', unsafe_allow_html=True)

            st.markdown("**정산서 교재별 판매 내역:**")
            st.dataframe(df_js.style.format({
                '정산서_판매수량': '{:,.0f}', '정산서_판매금액': '{:,.0f}'
            }), use_container_width=True, hide_index=True)

    # ── STEP 4: 결과 다운로드 ──
    st.markdown('<div class="step-header">💾 STEP 4. 결과 다운로드</div>', unsafe_allow_html=True)

    # 결과 엑셀 생성
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # FLOW 전체 데이터
        filtered.to_excel(writer, sheet_name='FLOW_필터결과', index=False)

        # 자체교재 집계
        if len(filtered_a) > 0:
            a_agg = filtered_a.groupby('교재명_정리').agg(
                판매건수=(amt_col, 'count'),
                총납입금액=(amt_col, 'sum'),
            ).sort_values('총납입금액', ascending=False).reset_index()
            a_agg.columns = ['교재명', '판매건수', '총납입금액']
            a_agg.to_excel(writer, sheet_name='자체교재_집계', index=False)

        # 시판교재 집계
        if len(filtered_c) > 0:
            c_agg = filtered_c.groupby('교재명_정리').agg(
                판매건수=(amt_col, 'count'),
                총납입금액=(amt_col, 'sum'),
            ).sort_values('총납입금액', ascending=False).reset_index()
            c_agg.columns = ['교재명', '판매건수', '총납입금액']
            c_agg.to_excel(writer, sheet_name='시판교재_집계', index=False)

        # 요약
        summary_data = {
            '항목': ['전체건수', '총납입금액', '자체교재(A)', '시판교재(C)', '환불건수'],
            '값': [total_count, total_all, total_a, total_c, refund_count]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='요약', index=False)

    output.seek(0)

    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button(
            "📥 분석 결과 엑셀 다운로드",
            data=output.getvalue(),
            file_name=f"{campus_input}_정산분석결과.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col_dl2:
        st.info(f"파일명: {campus_input}_정산분석결과.xlsx")

    st.markdown("")
    st.markdown('<div class="ok-box">✅ 분석 완료! 결과 엑셀을 다운로드하여 정산서에 반영하세요.</div>', unsafe_allow_html=True)

# ── Footer ──
st.markdown("---")
st.markdown("<p style='text-align:center;color:#aaa;font-size:12px;'>교재 정산 자동화 시스템 v1.0 | FLOW 데이터 기반 자동 추출</p>", unsafe_allow_html=True)
